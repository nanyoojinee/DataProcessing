import openpyxl

# 엑셀 파일 경로
excel_file_path = r'C:\Users\정유진\Desktop\[플리옥션]PMS데이터\플리옥션납품\품질점검.xlsx'

# 합친 결과를 저장할 시트 이름
output_sheet_name = 'CombinedData'

# 엑셀 파일 열기
workbook = openpyxl.load_workbook(excel_file_path)

# 결과를 저장할 새로운 시트 생성
output_sheet = workbook.create_sheet(output_sheet_name)

# 시트를 숫자로 시작하여 읽어서 합치기
for sheet_name in workbook.sheetnames:
    if sheet_name[0].isdigit():  # 시트 이름이 숫자로 시작하는 경우
        current_sheet = workbook[sheet_name]
        for row in current_sheet.iter_rows(min_row=1, max_row=current_sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                output_sheet.append([cell.value])

# 엑셀 파일 저장
workbook.save(excel_file_path)

print("숫자로 시작하는 시트가 하나의 시트로 합쳐졌습니다.")
