import openpyxl
import os

# 폴더 경로
folder_path = r"C:\Users\정유진\Desktop\납품_플리옥션\6.보도자료"

# 새로운 엑셀 파일명
output_file = r"C:\Users\정유진\Desktop\납품_플리옥션\6.보도자료\보도자료2.xlsx"

# 모든 값을 저장할 리스트
data = []

# 폴더 경로에 있는 모든 엑셀 파일을 반복
for filename in os.listdir(folder_path):
    if filename.endswith(".xlsx"):
        input_file = os.path.join(folder_path, filename)
        workbook = openpyxl.load_workbook(input_file)
        sheet = workbook.active

        # 원본 엑셀 파일의 B열부터 I열의 데이터를 가져와서 data에 추가
        for row in sheet.iter_rows(min_row=2, values_only=True):
            for cell_value in row[1:8]:  # B열부터 I열까지 데이터만 가져옵니다.
                if cell_value is not None:
                    data.append(cell_value)

# 새로운 엑셀 파일 생성 및 데이터 저장
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

# 데이터를 1열에 저장
for cell_value in data:
    new_sheet.append([cell_value])

# 새로운 엑셀 파일 저장
new_workbook.save(output_file)

print("새로운 엑셀 파일 생성 완료: " + output_file)
